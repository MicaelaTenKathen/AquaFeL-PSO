import openpyxl

from Data.limits import Limits
from Environment.plot import Plots
from Environment.map import Map
from Benchmark.benchmark_functions import Benchmark_function
from Environment.bounds import Bounds
from Data.utils import Utils
from Environment.contamination_areas import DetectContaminationAreas
from Environment.plot import Plots

from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import RBF
from sklearn.metrics import mean_squared_error

import numpy as np
import random
import math
import gym
import copy

from deap import base
from deap import creator
from deap import tools

import sys, os

sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

"""[https://deap.readthedocs.io/en/master/examples/pso_basic.html]"""


def createPart():
    """
    Creation of the objects "FitnessMax" and "Particle"
    """

    creator.create("FitnessMax", base.Fitness, weights=(1.0,))
    creator.create("Particle", np.ndarray, fitness=creator.FitnessMax, speed=None, smin=None, smax=None,
                   best=None)
    creator.create("BestGP", np.ndarray, fitness=creator.FitnessMax)


class PSOEnvironment(gym.Env):

    def __init__(self, resolution, ys, method, initial_seed, initial_position, vehicles=4, exploration_distance = 100,
                 exploitation_distance=200, reward_function='mse', behavioral_method=0, type_error='all_map',
                 stage="exploration"):
        self.type_error = type_error
        self.exploration_distance = exploitation_distance
        self.exploitation_distance = exploitation_distance
        self.stage = stage
        self.dist_pre = 0
        self.f = None
        self.k = None
        self.dict_ = {}
        self.max_bench = list()
        self.dict_impo_ = {}
        self.dict_bench = {}
        self.dict_coord_ = {}
        self.dict_sample_x = {}
        self.dict_error_peak = {}
        self.dict_sample_y = {}
        self.dict_fitness = {}
        self.dict_mu = {}
        self.dict_index = {}
        self.dict_sigma = {}
        self.dict_max_sigma = {}
        self.dict_max_mu = {}
        self.dict_global_best = {}
        self.dict_error = {}
        self.dict_centers = {}
        self.coord_centers = []
        self.max_centers_bench = []
        self.centers = 0
        self.vehicles = vehicles
        self.assig_centers = np.zeros((self.vehicles, 1))
        self.population = 4
        self.resolution = resolution
        self.smin = 0
        self.smax = 3
        self.size = 2
        self.wmin = 0.4 / (15000 / ys)
        self.wmax = 0.9 / (15000 / ys)
        self.xs = int(10000 / (15000 / ys))
        self.ys = ys
        self.max_peaks_bench = list()
        self.max_peaks_mu = list()
        ker = RBF(length_scale=10, length_scale_bounds=(1e-1, 10))
        self.gpr = GaussianProcessRegressor(kernel=ker, alpha=1e-6)  # optimizer=None)
        self.x_h = []
        self.y_h = []
        self.x_p = []
        self.y_p = []
        self.fitness = []
        self.y_data = []
        self.mu_max = 0
        self.x_bench = None
        self.y_bench = None
        self.n_plot = float(1)
        self.s_n = np.array([True, True, True, True])
        self.s_ant = np.zeros(4)
        self.samples = None
        self.dist_ant = None
        self.sigma_best = []
        self.mu_best = []
        self.action_zone = list()
        self.action_zone_index = list()
        self.n_data = 1
        self.num = 0
        self.save = 0
        self.num_of_peaks = 0
        self.save_dist = [25, 50, 75, 100, 125, 150, 175, 200]
        self.seed = initial_seed
        self.mu = []
        self.sigma = []
        self.g = 0
        self.post_array = np.array([1, 1, 1, 1])
        self.distances = np.zeros(4)
        self.distances_exploit = np.zeros(4)
        self.lam = 0.375
        self.part_ant = np.zeros((1, 8))
        self.part_ant_exploit = np.zeros((1, 8))
        self.last_sample, self.k, self.f, self.samples, self.ok = 0, 0, 0, 0, False
        self.error_data = []
        self.error_data1 = []
        self.it = []
        self.coordinate_bench_max = []
        self.bench_max = 0
        self.method = method
        self.error = []
        self.index_a = list()
        self.peaks = list()
        self.ERROR_data = []
        self.error_comparison = []
        self.error_comparison1 = []
        self.error_comparison2 = []
        self.error_comparison3 = []
        self.error_comparison4 = []
        self.error_comparison5 = []
        self.error_comparison6 = []
        self.error_comparison7 = []
        self.error_comparison8 = []
        self.bench_array = []
        self.error_distance = []
        self.duplicate = False
        self.array_part = np.zeros((1, 8))
        self.reward_function = reward_function
        self.behavioral_method = behavioral_method
        self.initial_position = initial_position
        if self.method == 0:
            self.state = np.zeros(22, )
        else:
            self.state = np.zeros((6, self.xs, self.ys))

        self.grid_or = Map(self.xs, ys).black_white()

        self.grid_min, self.grid_max, self.grid_max_x, self.grid_max_y = 0, self.ys, self.xs, self.ys

        self.p = 1
        self.max_peaks = None

        self.df_bounds, self.X_test = Bounds(self.resolution, self.xs, self.ys, load_file=False).map_bound()
        self.secure, self.df_bounds = Bounds(self.resolution, self.xs, self.ys).interest_area()

        # self.secure = navigation_map
        # print(self.secure)
        self.X_test_y = self.X_test[1]
        self.bench_function = None

        self.plot = Plots(self.xs, self.ys, self.X_test, self.secure, self.bench_function, self.grid_min, self.grid_or)

        # self.action_space = gym.spaces.Box(low=0.0, high=4.0, shape=(4,))
        # self.state_space = gym.spaces.Box()

        self.util = Utils()

        createPart()

    def generatePart(self):

        """
        Generates a random position and a random speed for the particles (drones).
        """
        part = creator.Particle([self.initial_position[self.p, i] for i in range(self.size)])
        part.speed = np.array([random.uniform(self.smin, self.smax) for _ in range(self.size)])
        part.smin = self.smin
        part.smax = self.smax
        self.p += 1

        return part

    def updateParticle_n(self, c1, c2, c3, c4, part):

        """
        Calculates the speed and the position of the particles (drones).
        """

        if self.behavioral_method == 0:
            u1 = np.array([random.uniform(0, c1) for _ in range(len(part))])
            u2 = np.array([random.uniform(0, c2) for _ in range(len(part))])
            u3 = np.array([random.uniform(0, c3) for _ in range(len(part))])
            u4 = np.array([random.uniform(0, c4) for _ in range(len(part))])
        else:
            u1 = c1
            u2 = c2
            u3 = c3
            u4 = c4

        v_u1 = u1 * (part.best - part)
        v_u2 = u2 * (self.best - part)
        v_u3 = u3 * (self.sigma_best - part)
        v_u4 = u4 * (self.mu_best - part)
        w = 1
        part.speed = v_u1 + v_u2 + v_u3 + v_u4 + part.speed * w
        for i, speed in enumerate(part.speed):
            if abs(speed) < part.smin:
                part.speed[i] = math.copysign(part.smin, speed)
            elif abs(speed) > part.smax:
                part.speed[i] = math.copysign(part.smax, speed)
        part[:] = part + part.speed

        return part

    def tool(self):

        """
        The operators are registered in the toolbox with their parameters.
        """
        self.toolbox = base.Toolbox()
        self.toolbox.register("particle", self.generatePart)
        self.toolbox.register("population", tools.initRepeat, list, self.toolbox.particle)
        self.toolbox.register("update", self.updateParticle_n)

        return self.toolbox

    def swarm(self):

        """
        Creates a population.
        """
        toolbox = self.tool()
        self.pop = toolbox.population(n=self.population)
        self.best = self.pop[0]

        return self.best, self.pop

    def statistic(self):

        """
        Visualizes the stats of the code.
        """

        self.stats = tools.Statistics(lambda ind: ind.fitness.values)
        self.stats.register("avg", np.mean)
        self.stats.register("std", np.std)
        self.stats.register("min", np.min)
        self.stats.register("max", np.max)

        self.logbook = tools.Logbook()
        self.logbook.header = ["gen", "evals"] + self.stats.fields

        return self.stats, self.logbook

    def reset(self):

        """
        Initialization of the pso.
        """
        self.reset_variables()
        self.bench_function, self.bench_array, self.num_of_peaks, self.index_a = Benchmark_function(self.grid_or,
                                                                                                    self.resolution,
                                                                                                    self.xs, self.ys,
                                                                                                    self.X_test,
                                                                                                    self.seed).create_new_map()

        self.max_contamination()
        self.generatePart()
        self.tool()
        print(self.bench_array)
        random.seed(self.seed)
        self.swarm()
        self.statistic()
        self.peaks_bench()
        self.detect_areas = DetectContaminationAreas(self.X_test, self.bench_array, vehicles=self.vehicles,
                                                     area=self.xs)
        self.max_peaks = self.detect_areas.real_peaks()
        print(self.max_peaks)
        self.state = self.first_values()
        return self.state

    def reset_variables(self):
        self.f = None
        self.k = None
        self.x_h = []
        self.y_h = []
        self.stage = "exploration"
        self.coord_centers = []
        self.max_centers_bench = []
        self.x_p = []
        self.y_p = []
        self.action_zone = list()
        self.max_bench = list()
        self.action_zone_index = list()
        self.fitness = []
        self.y_data = []
        self.dict_ = {}
        self.dict_impo_ = {}
        self.dict_error_peak = {}
        self.assig_centers = np.zeros((self.vehicles, 1))
        self.dict_coord_ = {}
        self.dict_sample_x = {}
        self.dict_error = {}
        self.dict_sample_y = {}
        self.dict_index = {}
        self.dict_fitness = {}
        self.dict_mu = {}
        self.dict_bench = {}
        self.dict_sigma = {}
        self.dict_centers = {}
        self.dict_max_sigma = {}
        self.dict_max_mu = {}
        self.dict_global_best = {}
        self.centers = 0
        self.x_bench = None
        self.y_bench = None
        self.n_plot = float(1)
        self.s_n = np.array([True, True, True, True])
        self.s_ant = np.zeros(4)
        self.samples = None
        self.dist_ant = None
        self.sigma_best = []
        self.mu_best = []
        self.coordinate_bench_max = []
        self.n_data = 1
        self.mu = []
        self.max_peaks_bench = list()
        self.max_peaks_mu = list()
        self.p = 0
        self.sigma = []
        self.post_array = np.array([1, 1, 1, 1])
        self.distances = np.zeros(4)
        self.part_ant = np.zeros((1, 8))
        self.distances_exploit = np.zeros(4)
        self.part_ant_exploit = np.zeros((1, 8))
        self.last_sample, self.k, self.f, self.samples, self.ok = 0, 0, 0, 0, False
        self.error_data = []
        self.save = 0
        self.error_comparison = []
        self.error_distance = []
        self.error = None

        self.it = []
        self.error = []
        self.duplicate = False
        self.array_part = np.zeros((1, 8))
        self.seed += 1

        if self.method == 0:
            self.state = np.zeros(22, )
        else:
            self.state = np.zeros((6, self.xs, self.ys))

        self.num += 1
        self.g = 0
        self.distances = np.zeros(4)

    def max_contamination(self):
        self.bench_max, self.coordinate_bench_max = self.obtain_max(self.bench_array)

    def pso_fitness(self, part):

        """
        Obtains the local best (part.best) of each particle (drone) and the global best (best) of the swarm (fleet).
        """

        part.fitness.values = self.new_fitness(part)

        if self.ok:
            self.check_duplicate(part)
        else:
            if part.best.fitness < part.fitness:
                part.best = creator.Particle(part)
                part.best.fitness.values = part.fitness.values
            if self.stage == "exploration":
                if self.best.fitness < part.fitness:
                    self.best = creator.Particle(part)
                    self.best.fitness.values = part.fitness.values

        return self.ok, part

    def new_fitness(self, part):
        part, self.s_n = Limits(self.secure, self.xs, self.ys).new_limit(self.g, part, self.s_n, self.n_data,
                                                                         self.s_ant, self.part_ant)
        self.x_bench = int(part[0])
        self.y_bench = int(part[1])

        new_fitness_value = [self.bench_function[self.x_bench][self.y_bench]]
        return new_fitness_value

    def gp_regression(self):

        """
        Fits the gaussian process.
        """

        x_a = np.array(self.x_h).reshape(-1, 1)
        y_a = np.array(self.y_h).reshape(-1, 1)
        x_train = np.concatenate([x_a, y_a], axis=1).reshape(-1, 2)
        y_train = np.array(self.fitness).reshape(-1, 1)

        self.gpr.fit(x_train, y_train)
        self.gpr.get_params()

        self.mu, self.sigma = self.gpr.predict(self.X_test, return_std=True)
        post_ls = np.min(np.exp(self.gpr.kernel_.theta[0]))
        r = self.n_data - 1
        self.post_array[r] = post_ls

        return self.post_array

    def sort_index(self, array, rev=True):
        index = range(len(array))
        s = sorted(index, reverse=rev, key=lambda i: array[i])
        return s

    def peaks_bench(self):
        for i in range(len(self.index_a)):
            self.max_peaks_bench.append(self.bench_array[round(self.index_a[i])])
            # print(round(self.index_a[i]))
        # self.peaks = self.sort_index(self.bench_array)[:self.num_of_peaks]
        # for i in range(len(self.peaks)):
        #   self.max_peaks_bench.append(self.bench_array[self.peaks[i]])

    def peaks_mu(self):
        self.max_peaks_mu = list()
        for i in range(len(self.index_a)):
            # print(round(self.index_a[i]))
            # print('in_here')
            self.max_peaks_mu.append(self.mu[round(self.index_a[i])])

    def sigma_max(self):

        """
        Returns the coordinates of the maximum uncertainty (sigma_best) and the maximum contamination (mu_best).
        """

        sigma_max, self.sigma_best = self.obtain_max(self.sigma)
        mu_max, self.mu_best = self.obtain_max(self.mu)

        return self.sigma_best, self.mu_best

    def calculate_reward(self):
        if self.reward_function == 'mse':
            reward = -self.error_data[-1]
        elif self.reward_function == 'inc_mse':
            reward = self.error_data[-2] - self.error_data[-1]
        return reward

    def check_duplicate(self, part):
        self.duplicate = False
        for i in range(len(self.x_h)):
            if self.x_h[i] == self.x_bench and self.y_h[i] == self.y_bench:
                self.duplicate = True
                break
            else:
                self.duplicate = False
        if self.duplicate:
            pass
        else:
            self.x_h.append(int(part[0]))
            self.y_h.append(int(part[1]))
            self.fitness.append(part.fitness.values)

    def obtain_max_explotaition(self, array_function, action_zone):
        max_value = np.max(array_function)
        index_1 = np.where(array_function == max_value)
        index_x1 = index_1[0]

        coord = self.dict_coord_["action_zone%s" % action_zone]
        index_x2 = index_x1[0]
        index_x = int(coord[index_x2][0])
        index_y = int(coord[index_x2][1])

        index_xy = [index_x, index_y]
        coordinate_max = np.array(index_xy)

        return max_value, coordinate_max

    def take_sample(self, part, action_zone):

        """
        Obtains the local best (part.best) of each particle (drone) and the global best (best) of the swarm (fleet).
        """

        part.fitness.values = self.new_fitness(part)
        x_bench = part[0]
        y_bench = part[1]
        duplicate = False
        x_l = copy.copy(self.dict_sample_x["action_zone%s" % action_zone])
        y_l = copy.copy(self.dict_sample_y["action_zone%s" % action_zone])
        print("x_h", x_l)
        print("y_h", y_l)
        fitness = copy.copy(self.dict_fitness["action_zone%s" % action_zone])
        print("fitness", fitness)
        index_action_zone = copy.copy(self.dict_index["action_zone%s" % action_zone])
        print(x_bench, y_bench)
        for i in range(len(x_l)):
            if x_l[i] == x_bench and y_l[i] == y_bench:
                duplicate = True
                fitness[i] = part.fitness.values
                break
            else:
                duplicate = False
        if duplicate:
            pass
        else:
            x_l.append(int(part[0]))
            y_l.append(int(part[1]))
            fitness.append(part.fitness.values)

        x_a = np.array(x_l).reshape(-1, 1)
        y_a = np.array(y_l).reshape(-1, 1)
        x_train = np.concatenate([x_a, y_a], axis=1).reshape(-1, 2)
        y_train = np.array(fitness).reshape(-1, 1)

        self.gpr.fit(x_train, y_train)
        self.gpr.get_params()

        mu, sigma = self.gpr.predict(self.X_test, return_std=True)
        post_ls = np.min(np.exp(self.gpr.kernel_.theta[0]))
        r = self.n_data - 1
        self.post_array[r] = post_ls

        mu_available = list()
        sigma_available = list()

        for i in range(len(index_action_zone)):
            mu_available.append(mu[index_action_zone[i]])
            sigma_available.append(sigma[index_action_zone[i]])

        sigma_max, sigma_best = self.obtain_max_explotaition(sigma_available, action_zone)
        mu_max, mu_best = self.obtain_max_explotaition(mu_available, action_zone)

        self.dict_sample_x["action_zone%s" % action_zone] = copy.copy(x_l)
        self.dict_sample_y["action_zone%s" % action_zone] = copy.copy(y_l)
        self.dict_fitness["action_zone%s" % action_zone] = copy.copy(fitness)
        self.dict_mu["action_zone%s" % action_zone] = copy.copy(mu)
        self.dict_sigma["action_zone%s" % action_zone] = copy.copy(sigma)
        self.dict_max_sigma["action_zone%s" % action_zone] = copy.copy(sigma_best)
        self.dict_max_mu["action_zone%s" % action_zone] = copy.copy(mu_best)

    def first_values(self):

        """
        The output "out" of the method "initcode" is the positions of the particles (drones) after the first update of the
        gaussian process (initial state).
        method = 0 -> out = scalar vector
        out = [px_1, py_1, px_2, py_2, px_3, py_3, px_4, py_4, lbx_1, lby_1, lbx_2, lby_2, lbx_3, lby_3, lbx_4, lby_4, gbx, gby,
               sbx, sgy, mbx, mby]
               where:
               px: x coordinate of the drone position
               py: y coordinate of the drone position
               lbx: x coordinate of the local best
               lby: y coordinate of the local best
               gbx: x coordinate of the global best
               gby: y coordinate of the global best
               sbx: x coordinate of the sigma best (maximum uncertainty)
               sby: y coordinate of the sigma best (maximum uncertainty)
               mbx: x coordinate of the mean best (maximum contamination)
               mby: y coordinate of the mean best (maximum contamination)
        method = 1 -> out = images
        :param c1: weight that determinate the importance of the local best component
        :param c2: weight that determinate the importance of the global best component
        :param c3: weight that determinate the importance of the maximum uncertainty component
        :param c4: weight that determinate the importance of the maximum contamination component
        :param lam: ratio of one of the different length scales [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        :param post_array: refers to the posterior length scale of the surrogate model [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        """

        for part in self.pop:

            part.fitness.values = self.new_fitness(part)

            if self.n_plot > 4:
                self.n_plot = float(1)

            part.best = creator.Particle(part)
            part.best.fitness.values = part.fitness.values

            if self.best.fitness < part.fitness:
                self.best = creator.Particle(part)
                self.best.fitness.values = part.fitness.values

            self.part_ant, self.distances = self.util.distance_part(self.g, self.n_data, part, self.part_ant,
                                                                    self.distances, self.array_part, dfirst=True)

            self.check_duplicate(part)

            self.post_array = self.gp_regression()

            self.samples += 1

            self.n_data += 1
            if self.n_data > 4:
                self.n_data = 1

        # self.MSE_data1, self.it = self.util.mse(self.g, self.bench_array, self.mu)
        self.error = self.calculate_error()
        self.error_data.append(self.error)
        self.it.append(self.g)

        self.sigma_best, self.mu_best = self.sigma_max()

        self.return_state()

        self.k = 4
        self.ok = False

        return self.state

    def save_data(self):
        mult = self.save_dist[self.save]
        mult_min = mult - 5
        mult_max = mult + 5
        # print(mult, mult_min, mult_max, np.max(self.distances))
        if mult_min <= np.max(self.distances) < mult_max:
            self.ERROR_data = self.calculate_error()
            # print(np.max(self.distances))
            if self.save == 0:
                self.error_comparison1.append(self.ERROR_data)
            elif self.save == 1:
                self.error_comparison2.append(self.ERROR_data)
            elif self.save == 2:
                self.error_comparison3.append(self.ERROR_data)
            elif self.save == 3:
                self.error_comparison4.append(self.ERROR_data)
            elif self.save == 4:
                self.error_comparison5.append(self.ERROR_data)
            elif self.save == 5:
                self.error_comparison6.append(self.ERROR_data)
            elif self.save == 6:
                self.error_comparison7.append(self.ERROR_data)
            self.save += 1

    def obtain_max(self, array_function):
        max_value = np.max(array_function)
        index_1 = np.where(array_function == max_value)
        index_x1 = index_1[0]

        index_x2 = index_x1[0]
        index_x = int(self.X_test[index_x2][0])
        index_y = int(self.X_test[index_x2][1])

        index_xy = [index_x, index_y]
        coordinate_max = np.array(index_xy)

        return max_value, coordinate_max

    def calculate_error(self):
        if self.type_error == 'all_map':
            self.error = mean_squared_error(y_true=self.bench_array, y_pred=self.mu)
        elif self.type_error == 'peaks':
            print(self.max_bench)
            print(self.max_centers_bench)
            for i in range(len(self.dict_centers)):
                coord = self.max_centers_bench[i]
                for j in range(len(self.X_test)):
                    coord_xtest = self.X_test[j]
                    if coord[0] == coord_xtest[0] and coord[1] == coord_xtest[1]:
                        max_az = self.dict_mu["action_zone%s" % i][j]
                        print(max_az)
                        break
                self.dict_error_peak["action_zone%s" % i] = abs(self.max_bench[i] - max_az)
        elif self.type_error == 'contamination_1':
            index_mu_max = [i for i in range(len(self.X_test)) if (self.X_test[i] == self.coordinate_bench_max).all()]
            index_mu_max = index_mu_max[0]
            # for i in range(len(self.X_test)):
            #     di = self.X_test[i]
            #     dix = di[0]
            #     diy = di[1]
            #     if dix == self.coordinate_bench_max[0] and diy == self.coordinate_bench_max[1]:
            #         mu_max = self.mu[i]
            #         break
            mu_max = self.mu[index_mu_max]
            mu_max = mu_max[0]
            self.error = self.bench_max - mu_max
        elif self.type_error == 'contamination':
            self.peaks_mu()
            self.error = mean_squared_error(y_true=self.max_peaks_bench, y_pred=self.max_peaks_mu)
        elif self.type_error == 'action_zone':
            estimated_all = list()
            for i in range(len(self.coord_centers)):
                bench_action = copy.copy(self.dict_bench["action_zone%s" % i])
                estimated_action = list()
                index_action = copy.copy(self.dict_index["action_zone%s" % i])
                mu_action = copy.copy(self.dict_mu["action_zone%s" % i])
                for j in range(len(index_action)):
                    estimated_action.append(mu_action[index_action[j]])
                    estimated_all.append(mu_action[index_action[j]])
                error_action = mean_squared_error(y_true=bench_action, y_pred=estimated_action)
                self.dict_error["action_zone%s" % i] = copy.copy(error_action)
            self.error = mean_squared_error(y_true=self.action_zone, y_pred=estimated_all)
        return self.error

    def allocate_vehicles(self):
        center_zone = copy.copy(self.coord_centers)
        population = copy.copy(self.pop)
        num = 0
        asvs = np.arange(0, self.vehicles, 1)
        elements_repeat_all = math.trunc(self.vehicles / self.centers)
        elements_repeat = self.vehicles - elements_repeat_all * self.centers
        repeat_all = np.full((self.centers, 1), elements_repeat_all)
        repeat = np.zeros((self.centers, 1))
        while num < elements_repeat:
            repeat[num] = 1
            num += 1
        for i in range(len(repeat_all)):
            assig_vehicles = list()
            z = repeat_all[i] + repeat[i]
            o = 0
            while o < z:
                asv = 0
                for part in population:
                    if asv == 0:
                        low = math.sqrt((center_zone[i, 0] - part[0]) ** 2 + (center_zone[i, 1] - part[1]) ** 2)
                        index = 0
                    else:
                        dista = math.sqrt((center_zone[i, 0] - part[0]) ** 2 + (center_zone[i, 1] - part[1]) ** 2)
                        if dista < low:
                            low = copy.copy(dista)
                            index = copy.copy(asv)
                    asv += 1
                assig_vehicles.append(asvs[index])
                self.assig_centers[asvs[index]] = i
                del population[index]
                asvs = np.delete(asvs, index)
                o += 1
            self.dict_centers["action_zone%s" % i] = assig_vehicles
        print(self.dict_centers)
        zones = 0
        while zones < self.centers:
            self.dict_max_mu["action_zone%s" % zones] = center_zone[zones]
            print(self.dict_max_mu)
            self.dict_max_sigma["action_zone%s" % zones] = self.sigma_best
            self.dict_sample_x["action_zone%s" % zones] = self.x_h
            self.dict_sample_y["action_zone%s" % zones] = self.y_h
            self.dict_fitness["action_zone%s" % zones] = self.fitness
            zones += 1

    def step_stage_exploration(self, action):

        """
        The output "out" of the method "step" is the positions of the particles (drones) after traveling 1000 m
        (scaled).

        method = 0 -> out = scalar vector
        out = [px_1, py_1, px_2, py_2, px_3, py_3, px_4, py_4, lbx_1, lby_1, lbx_2, lby_2, lbx_3, lby_3, lbx_4, lby_4,
               gbx, gby, sbx, sgy, mbx, mby]
               where:
               px: x coordinate of the drone position
               py: y coordinate of the drone position
               lbx: x coordinate of the local best
               lby: y coordinate of the local best
               gbx: x coordinate of the global best
               gby: y coordinate of the global best
               sbx: x coordinate of the sigma best (maximum uncertainty)
               sby: y coordinate of the sigma best (maximum uncertainty)
               mbx: x coordinate of the mean best (maximum contamination)
               mby: y coordinate of the mean best (maximum contamination)

        method = 1 -> out = images

        :param c1: weight that determinate the importance of the local best component
        :param c2: weight that determinate the importance of the global best component
        :param c3: weight that determinate the importance of the maximum uncertainty component
        :param c4: weight that determinate the importance of the maximum contamination component
        :param lam: ratio of one of the different length scales [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        :param post_array: refers to the posterior length scale of the surrogate model [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        """
        dis_steps = 0
        dist_ant = np.mean(self.distances)
        self.dist_pre = np.max(self.distances)
        self.n_data = 1
        self.f += 1

        while dis_steps < 10:

            previous_dist = np.max(self.distances)

            for part in self.pop:
                self.toolbox.update(action[0], action[1], action[2], action[3], part)

            for part in self.pop:
                self.ok, part = self.pso_fitness(part)
                self.part_ant, self.distances = self.util.distance_part(self.g, self.n_data, part, self.part_ant,
                                                                        self.distances, self.array_part, dfirst=False)

                self.n_data += 1
                if self.n_data > 4:
                    self.n_data = 1

            if (np.mean(self.distances) - self.last_sample) >= (np.min(self.post_array) * self.lam):
                self.k += 1
                self.ok = True
                self.last_sample = np.mean(self.distances)

                for part in self.pop:
                    self.ok, part = self.pso_fitness(part)

                    self.post_array = self.gp_regression()

                    self.samples += 1

                    self.n_data += 1
                    if self.n_data > 4:
                        self.n_data = 1

                self.error = self.calculate_error()
                self.error_data.append(self.error)
                self.it.append(self.g)

                self.sigma_best, self.mu_best = self.sigma_max()

                self.ok = False

            dis_steps = np.mean(self.distances) - dist_ant
            if np.max(self.distances) == previous_dist:
                break

            self.save_data()
            self.g += 1

        self.return_state()

        reward = self.calculate_reward()

        self.logbook.record(gen=self.g, evals=len(self.pop), **self.stats.compile(self.pop))

        if (self.distances >= self.exploration_distance).any() or np.max(self.distances) == self.dist_pre:
            done = False
            self.dict_, self.dict_coord_, self.dict_impo_, self.centers, self.coord_centers, self.dict_index, \
            self.dict_bench, self.action_zone, self.max_centers_bench, self.max_bench = self.detect_areas.areas_levels(self.mu)
            self.plot.action_areas(self.dict_coord_, self.dict_impo_, self.centers)
            self.allocate_vehicles()
            self.obtain_global()
           # for part in self.pop:
            #    self.part_ant_exploit, self.distances_exploit = self.util.distance_part(self.g, self.n_data, part,
             #                                                                           self.part_ant,
              #                                                                          self.distances, self.array_part,
               #                                                                         dfirst=True)
        else:
            done = False
        return self.state, reward, done, {}

    def obtain_global(self):
        for i in range(len(self.dict_centers)):
            list_vehicle = self.dict_centers["action_zone%s" % i]
            for j in range(len(list_vehicle)):
                part = self.pop[list_vehicle[j]]
                if j == 0:
                    best = part.best
                    best.fitness.values = part.fitness.values
                else:
                    if best.fitness < part.fitness:
                        best = part.best
                        best.fitness.values = part.fitness.values
            self.dict_global_best["action_zone%s" % i] = best

    def step_stage_exploitation(self, action):

        """
        The output "out" of the method "step" is the positions of the particles (drones) after traveling 1000 m
        (scaled).

        method = 0 -> out = scalar vector
        out = [px_1, py_1, px_2, py_2, px_3, py_3, px_4, py_4, lbx_1, lby_1, lbx_2, lby_2, lbx_3, lby_3, lbx_4, lby_4,
               gbx, gby, sbx, sgy, mbx, mby]
               where:
               px: x coordinate of the drone position
               py: y coordinate of the drone position
               lbx: x coordinate of the local best
               lby: y coordinate of the local best
               gbx: x coordinate of the global best
               gby: y coordinate of the global best
               sbx: x coordinate of the sigma best (maximum uncertainty)
               sby: y coordinate of the sigma best (maximum uncertainty)
               mbx: x coordinate of the mean best (maximum contamination)
               mby: y coordinate of the mean best (maximum contamination)

        method = 1 -> out = images

        :param c1: weight that determinate the importance of the local best component
        :param c2: weight that determinate the importance of the global best component
        :param c3: weight that determinate the importance of the maximum uncertainty component
        :param c4: weight that determinate the importance of the maximum contamination component
        :param lam: ratio of one of the different length scales [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        :param post_array: refers to the posterior length scale of the surrogate model [Equation 7
        (https://doi.org/10.3390/electronics10131605)]
        """
        dis_steps = 0
        dist_ant = np.mean(self.distances)
        self.dist_pre = np.max(self.distances)
        self.n_data = 1
        self.f += 1

        while dis_steps < 10:

            previous_dist = np.max(self.distances)

            asv = 0
            for part in self.pop:
                action_zone = int(self.assig_centers[asv])
                self.mu_best = self.dict_max_mu["action_zone%s" % action_zone]
                self.sigma_best = self.dict_max_sigma["action_zone%s" % action_zone]
                self.best = self.dict_global_best["action_zone%s" % action_zone]
                print("local", part.best, "global", self.best, "mu", self.mu_best)
                self.toolbox.update(action[0], action[1], action[2], action[3], part)
                asv += 1

            for part in self.pop:
                self.ok, part = self.pso_fitness(part)

            self.obtain_global()

            for part in self.pop:
                self.part_ant, self.distances = self.util.distance_part(self.g, self.n_data, part, self.part_ant,
                                                                        self.distances, self.array_part, dfirst=False)

                #self.part_ant_exploit, self.distances_exploit = self.util.distance_part(self.g, self.n_data, part,
                 #                                                                       self.part_ant,
                  #                                                                      self.distances, self.array_part,
                   #                                                                     dfirst=False)

                self.n_data += 1
                if self.n_data > 4:
                    self.n_data = 1

            if (np.mean(self.distances) - self.last_sample) >= (np.min(self.post_array) * self.lam):
                self.k += 1
                self.last_sample = np.mean(self.distances)

                for i in range(len(self.dict_centers)):
                    list_vehicle = self.dict_centers["action_zone%s" % i]
                    for j in range(len(list_vehicle)):
                        part = self.pop[list_vehicle[j]]
                        self.take_sample(part, i)

                        self.n_data += 1
                        if self.n_data > 4:
                            self.n_data = 1

                        self.samples += 1

                self.type_error = 'action_zone'
                self.error = self.calculate_error()
                self.error_data.append(self.error)
                self.it.append(self.g)
                self.type_error = 'peaks'
                self.error = self.calculate_error()

            dis_steps = np.mean(self.distances) - dist_ant
            if np.max(self.distances) == previous_dist:
                break

            # self.save_data()
            self.g += 1

        self.return_state()

        reward = self.calculate_reward()

        self.logbook.record(gen=self.g, evals=len(self.pop), **self.stats.compile(self.pop))

        if (self.distances >= self.exploitation_distance).any() or np.max(self.distances) == self.dist_pre:
            done = True
            print(self.dict_error)
            print(self.dict_error_peak)
        else:
            done = False
        return self.state, reward, done, {}

    def return_state(self):
        z = 0
        for part in self.pop:
            self.state = self.state
            if self.method == 0:
                self.state[z] = part[0]
                z += 1
                self.state[z] = part[1]
                z += 1
                self.state[z + 6] = part.best[0]
                self.state[z + 7] = part.best[1]
                if self.n_data == 4:
                    self.state[16] = self.best[0]
                    self.state[17] = self.best[1]
                    self.state[18] = self.sigma_best[0]
                    self.state[19] = self.sigma_best[1]
                    self.state[20] = self.mu_best[0]
                    self.state[21] = self.mu_best[1]
            else:
                posx = 2 * z
                posy = (2 * z) + 1
                self.state = self.plot.part_position(self.part_ant[:, posx], self.part_ant[:, posy], self.state, z)
                z += 1
                if self.n_data == 4:
                    self.state = self.plot.state_sigma_mu(self.mu, self.sigma, self.state)
            self.n_data += 1
            if self.n_data > 4:
                self.n_data = 1

    def step(self, action):
        if self.stage == "exploration":
            action = np.array([2.0187, 0, 3.2697, 0])
            self.state, reward, done, dic = self.step_stage_exploration(action)
            if (self.distances >= self.exploration_distance).any() or np.max(self.distances) == self.dist_pre:
                self.stage = "exploitation"
        elif self.stage == "exploitation":
            action = np.array([3.6845, 1.5614, 0, 3.1262])
            self.state, reward, done, dic = self.step_stage_exploitation(action)
        return self.state, reward, done, {}

    def data_out(self):

        """
        Return the first and the last position of the particles (drones).
        """

        return self.X_test, self.secure, self.bench_function, self.grid_min, self.sigma, \
               self.mu, self.error_data, self.it, self.part_ant, self.bench_array, self.grid_or, self.bench_max, \
               self.dict_mu, self.dict_sigma, self.centers, self.part_ant_exploit, self.dict_centers

    def error_value(self):
        return self.error_data

    def return_seed(self):
        return self.seed

    def distances_data(self):
        return self.distances

    def save_excel(self):
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.append(self.error_comparison1)
        wb.save('../Test/Chapter/Epsilon/ALLCONError_25.xlsx')

        wb2 = openpyxl.Workbook()
        hoja2 = wb2.active
        hoja2.append(self.error_comparison2)
        wb2.save('../Test/Chapter/Epsilon/ALLCONError_50.xlsx')

        wb3 = openpyxl.Workbook()
        hoja3 = wb3.active
        hoja3.append(self.error_comparison3)
        wb3.save('../Test/Chapter/Epsilon/ALLCONError_75.xlsx')

        wb4 = openpyxl.Workbook()
        hoja4 = wb4.active
        hoja4.append(self.error_comparison4)
        wb4.save('../Test/Chapter/Epsilon/ALLCONError_100.xlsx')

        wb5 = openpyxl.Workbook()
        hoja5 = wb5.active
        hoja5.append(self.error_comparison5)
        wb5.save('../Test/Chapter/Epsilon/ALLCONError_125.xlsx')

        wb6 = openpyxl.Workbook()
        hoja6 = wb6.active
        hoja6.append(self.error_comparison6)
        wb6.save('../Test/Chapter/Epsilon/ALLCONError_150.xlsx')

        wb7 = openpyxl.Workbook()
        hoja7 = wb7.active
        hoja7.append(self.error_comparison7)
        wb7.save('../Test/Chapter/Epsilon/ALLCONErrorE_175.xlsx')

import math
import numpy as np
import openpyxl


class Utils:
    def distance_part(self, g, n_data, part, part_ant, distances, init=False):
        if init:
            if n_data == 1:
                part_ant[0, 0] = part[0]
                part_ant[0, 1] = part[1]
            if n_data == 2:
                part_ant[0, 2] = part[0]
                part_ant[0, 3] = part[1]
            if n_data == 3:
                part_ant[0, 4] = part[0]
                part_ant[0, 5] = part[1]
            if n_data == 4:
                part_ant[0, 6] = part[0]
                part_ant[0, 7] = part[1]
        else:
            if n_data == 1:
                part_ant[g + 1, 0] = part[0]
                part_ant[g + 1, 1] = part[1]
                distances[0] = math.sqrt(
                    (part_ant[g + 1, 0] - part_ant[g, 0]) ** 2 + (part_ant[g + 1, 1] - part_ant[g, 1])
                    ** 2) + distances[0]
            elif n_data == 2:
                part_ant[g + 1, 2] = part[0]
                part_ant[g + 1, 3] = part[1]
                distances[1] = math.sqrt(
                    (part_ant[g + 1, 2] - part_ant[g, 2]) ** 2 + (part_ant[g + 1, 3] - part_ant[g, 3])
                    ** 2) + distances[1]
            elif n_data == 3:
                part_ant[g + 1, 4] = part[0]
                part_ant[g + 1, 5] = part[1]
                distances[2] = math.sqrt(
                    (part_ant[g + 1, 4] - part_ant[g, 4]) ** 2 + (part_ant[g + 1, 5] - part_ant[g, 5])
                    ** 2) + distances[2]
            elif n_data == 4:
                part_ant[g + 1, 6] = part[0]
                part_ant[g + 1, 7] = part[1]
                distances[3] = math.sqrt(
                    (part_ant[g + 1, 6] - part_ant[g, 6]) ** 2 + (part_ant[g + 1, 7] - part_ant[g, 7])
                    ** 2) + distances[3]

        return part_ant, distances

    def mse(self, g, y_data, mu_data, samples, MSE_data, it):
        total_suma = 0
        y_array = np.array(y_data)
        mu_array = np.array(mu_data)
        for i in range(len(mu_array)):
            total_suma = (float(y_array[i]) - float(mu_array[i])) ** 2 + total_suma
        MSE = total_suma / samples
        MSE_data.append(MSE)
        it.append(g)
        return MSE_data, it

    def savexlsx(self, MSE_data, sigma_data, mu_data, distances, it, seed):
        for i in range(len(mu_data)):
            mu_data[i] = float(mu_data[i])

        wb1 = openpyxl.Workbook()
        hoja1 = wb1.active
        hoja1.append(MSE_data)
        wb1.save('Pruebas/Error' + str(seed[0]) + '.xlsx')

        wb2 = openpyxl.Workbook()
        hoja2 = wb2.active
        hoja2.append(sigma_data)
        wb2.save('Pruebas/Sigma' + str(seed[0]) + '.xlsx')

        wb3 = openpyxl.Workbook()
        hoja3 = wb3.active
        hoja3.append(mu_data)
        wb3.save('Pruebas/Mu' + str(seed[0]) + '.xlsx')

        wb4 = openpyxl.Workbook()
        hoja4 = wb4.active
        hoja4.append(list(distances))
        wb4.save('Pruebas/Distance' + str(seed[0]) + '.xlsx')

        wb5 = openpyxl.Workbook()
        hoja5 = wb5.active
        hoja5.append(it)
        wb5.save('Pruebas/Data' + str(seed[0]) + '.xlsx')
